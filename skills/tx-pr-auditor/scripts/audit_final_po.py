#!/usr/bin/env python3
"""
TX PR Auditor - Final PO validation baseline.

This script validates submitted Final PO rows against EPMS facts and the PR
Model. It intentionally follows the skill architecture as a staged pipeline and
fails closed when the current rule inputs are not sufficient for a confident
decision.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, replace
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple


PLANNING_ITEM_CODE = "350001000403"
OPERATION_ITEM_CODE = "350000592793"
ANTENNA_ITEM_CODES = {
    "350001095405": (0.0, 0.6),
    "350001095406": (0.6, 1.2),
    "350001095407": (1.2, 1.8),
    "350001095408": (1.8, None),
}
FINAL_PO_SHEET_NAME = "条目明细"
EPMS_SHEET_NAME = "data"

FINAL_PO_FIELD_MAP = {
    "派工日期": "dispatch_date",
    "派工单号": "dispatch_order_number",
    "PO行号": "po_line_number",
    "需求单号": "request_number",
    "项目名称": "project_name",
    "项目编码": "project_code",
    "业务大类": "business_domain",
    "施工区域": "region",
    "采购区域": "purchasing_area",
    "分包商": "submitted_subcontractor",
    "逻辑站点名称": "logical_site_name",
    "逻辑站点编码": "du",
    "物理站点名称": "physical_site_name",
    "物理站点编码": "site_code",
    "外包代码": "submitted_item_code",
    "代码名称": "submitted_item_description",
    "量纲": "submitted_unit",
    "派工数量": "submitted_quantity",
    "结算数量": "settlement_quantity",
    "支付数量": "paid_quantity",
    "产品型号_备注": "product_model_remark",
    "派工单状态": "dispatch_status",
    "外包商编码": "subcontractor_code",
}

EPMS_FIELD_MAP = {
    "customer site code": "site_code",
    "customer site name": "site_name",
    "du code": "du",
    "region": "epms_region",
    "Province/State": "province_state",
    "Latitude (North Plus South Minus)": "latitude",
    "Longitude (East Plus West Minus)": "longitude",
    "TX Upgrade Scope": "tx_upgrade_scope",
    "BOQ Configuration": "boq_configuration",
    "Tx SOW": "tx_sow",
    "TX SOW Details": "tx_sow_details",
    "NE SOW Details": "ne_sow_details",
    "FE SOW Details": "fe_sow_details",
    "MW Config Antenna Size NE": "antenna_size_ne",
    "MW Config Antenna Size FE": "antenna_size_fe",
    "SubCon - TSS Team": "expected_tss_subcontractor",
    "Subcon PR - TSS": "existing_tss_pr",
    "SubCon - TI Team": "expected_ti_subcontractor",
    "Subcon PR - TI": "existing_ti_pr",
    "TX Cutover Date": "tx_cutover_date",
    "Subcon - Planning": "expected_planning_subcontractor",
    "Subcon PR - Planning": "existing_planning_pr",
}


@dataclass(frozen=True)
class RawDataset:
    final_po_rows: List[Dict[str, Any]]
    epms_rows: List[Dict[str, Any]]
    pr_model_rows: List[Dict[str, Any]]
    metadata: Dict[str, Any]


@dataclass(frozen=True)
class FinalPORecord:
    source_row: int
    raw: Dict[str, Any]
    canonical: Dict[str, Any]


@dataclass(frozen=True)
class EPMSRecord:
    source_row: int
    raw: Dict[str, Any]
    canonical: Dict[str, Any]


@dataclass(frozen=True)
class CanonicalDataset:
    final_po_records: List[FinalPORecord]
    epms_records: List[EPMSRecord]
    pr_model_rows: List[Dict[str, Any]]
    metadata: Dict[str, Any]


@dataclass(frozen=True)
class BusinessRecord:
    final_po: FinalPORecord
    epms: Optional[EPMSRecord]
    scope: str
    expected_subcontractor: str
    business_facts: Dict[str, Any]


@dataclass(frozen=True)
class BusinessDataset:
    records: List[BusinessRecord]
    pr_model_rows: List[Dict[str, Any]]
    metadata: Dict[str, Any]


@dataclass(frozen=True)
class ExpectedItem:
    code: str
    quantity: float
    scope: str
    reason: str


@dataclass(frozen=True)
class ExpectedRecord:
    business_record: BusinessRecord
    expected_items: List[ExpectedItem]
    expected_subcontractor: str
    model_evidence: str


@dataclass(frozen=True)
class ExpectedDataset:
    records: List[ExpectedRecord]
    metadata: Dict[str, Any]


@dataclass(frozen=True)
class AuditResult:
    final_po: FinalPORecord
    scope: str
    classification: str
    reason_code: str
    expected_items: List[ExpectedItem]
    expected_subcontractor: str
    expected_quantity: float
    normal_quantity: float
    duplicate_quantity: float
    explanation: str
    epms_evidence: str
    pr_model_evidence: str
    consumes_quantity: bool


@dataclass(frozen=True)
class AuditDataset:
    results: List[AuditResult]
    metadata: Dict[str, Any]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Validate Final PO.xlsx against EPMS.xlsx and pr_model.xlsx."
    )
    parser.add_argument(
        "--final-po",
        default="input/Final PO.xlsx",
        help="Path to Final PO.xlsx",
    )
    parser.add_argument(
        "--final-po-sheet",
        default=FINAL_PO_SHEET_NAME,
        help="Worksheet name to read from Final PO.xlsx",
    )
    parser.add_argument(
        "--epms",
        default="input/EPMS.xlsx",
        help="Path to EPMS.xlsx",
    )
    parser.add_argument(
        "--epms-sheet",
        default=EPMS_SHEET_NAME,
        help="Worksheet name to read from EPMS.xlsx",
    )
    parser.add_argument(
        "--pr-model",
        default="input/pr_model.xlsx",
        help="Path to pr_model.xlsx",
    )
    parser.add_argument(
        "--output",
        default="output/PR_Audit_Result.xlsx",
        help="Path for PR_Audit_Result.xlsx",
    )
    parser.add_argument(
        "--summary-json",
        help="Optional path for a JSON summary",
    )
    return parser.parse_args()


def require_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill
    except ModuleNotFoundError as exc:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `python3 -m pip install -r requirements.txt`."
        ) from exc
    return Workbook, load_workbook, Font, PatternFill


def require_file(path: Path, label: str) -> Path:
    if not path.exists():
        raise FileNotFoundError(f"{label} not found: {path}")
    return path


def normalize_header(value: Any) -> str:
    return str(value or "").strip()


def unique_headers(headers: Sequence[Any]) -> List[str]:
    counts: Dict[str, int] = defaultdict(int)
    out: List[str] = []
    for header in headers:
        name = normalize_header(header)
        if not name:
            name = "EMPTY"
        counts[name] += 1
        if counts[name] == 1:
            out.append(name)
        else:
            out.append(f"{name}__{counts[name]}")
    return out


def read_table(path: Path, sheet_name: str, header_row: int) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    _, load_workbook, _, _ = require_openpyxl()
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Required worksheet '{sheet_name}' not found in {path}. "
            f"Available sheets: {', '.join(wb.sheetnames)}"
        )
    ws = wb[sheet_name]
    header_values = [cell.value for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row))]
    headers = unique_headers(header_values)
    rows: List[Dict[str, Any]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        if not any(value not in (None, "") for value in row):
            continue
        record = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        record["_source_row"] = row_idx
        rows.append(record)
    return rows, {
        "path": str(path),
        "sheet": ws.title,
        "header_row": header_row,
        "row_count": len(rows),
        "column_count": len(headers),
    }


def read_pr_model(path: Path) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    _, load_workbook, _, _ = require_openpyxl()
    wb = load_workbook(path, read_only=True, data_only=True)
    target_sheet = "TX Line Item (After 21-Apr 26)"
    sheet_name = target_sheet if target_sheet in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows: List[Dict[str, Any]] = []
    section = ""
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        values = list(row[:8])
        first = text(values[0])
        if "TSS Model" in first:
            section = "TSS"
            continue
        if "TI Model" in first:
            section = "TI"
            continue
        if section and first and text(values[1]) and text(values[2]):
            rows.append(
                {
                    "source_row": row_idx,
                    "section": section,
                    "sow": first,
                    "code": text(values[1]),
                    "description": text(values[2]),
                    "unit": text(values[3]) or "Hop",
                    "quantity": to_float(values[4], default=1.0),
                    "rules": text(values[5]),
                    "is_mandatory": "Mandatory" in text(values[5]),
                    "worksheet": sheet_name,
                }
            )
    return rows, {
        "path": str(path),
        "sheet": sheet_name,
        "row_count": len(rows),
    }


def workbook_reader(
    final_po: Path,
    epms: Path,
    pr_model: Path,
    final_po_sheet: str = FINAL_PO_SHEET_NAME,
    epms_sheet: str = EPMS_SHEET_NAME,
) -> RawDataset:
    final_po_rows, final_meta = read_table(final_po, sheet_name=final_po_sheet, header_row=1)
    epms_rows, epms_meta = read_table(epms, sheet_name=epms_sheet, header_row=4)
    pr_rows, pr_meta = read_pr_model(pr_model)
    return RawDataset(
        final_po_rows=final_po_rows,
        epms_rows=epms_rows,
        pr_model_rows=pr_rows,
        metadata={
            "final_po": final_meta,
            "epms": epms_meta,
            "pr_model": pr_meta,
        },
    )


def canonicalize(raw: Dict[str, Any], field_map: Dict[str, str]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for source_field, canonical_field in field_map.items():
        out[canonical_field] = raw.get(source_field)
    return out


def field_mapper(raw: RawDataset) -> CanonicalDataset:
    final_records = [
        FinalPORecord(
            source_row=int(row["_source_row"]),
            raw=row,
            canonical=canonicalize(row, FINAL_PO_FIELD_MAP),
        )
        for row in raw.final_po_rows
    ]
    epms_records = [
        EPMSRecord(
            source_row=int(row["_source_row"]),
            raw=row,
            canonical=canonicalize(row, EPMS_FIELD_MAP),
        )
        for row in raw.epms_rows
    ]
    return CanonicalDataset(
        final_po_records=final_records,
        epms_records=epms_records,
        pr_model_rows=raw.pr_model_rows,
        metadata=raw.metadata,
    )


def canonical_builder(canonical: CanonicalDataset) -> CanonicalDataset:
    final_records = []
    for record in canonical.final_po_records:
        data = dict(record.canonical)
        data["du"] = text(data.get("du"))
        data["submitted_item_code"] = text(data.get("submitted_item_code"))
        data["submitted_quantity"] = to_float(data.get("submitted_quantity"), default=0.0)
        data["settlement_quantity"] = to_float(data.get("settlement_quantity"), default=0.0)
        data["dispatch_sort_key"] = dispatch_sort_key(data)
        final_records.append(replace(record, canonical=data))

    epms_records = []
    for record in canonical.epms_records:
        data = dict(record.canonical)
        data["du"] = text(data.get("du"))
        data["integration_end_date"] = first_nonempty(
            [data.get("tx_cutover_date")]
            + [value for key, value in record.raw.items() if key.startswith("actual end time")]
        )
        epms_records.append(replace(record, canonical=data))

    return CanonicalDataset(
        final_po_records=final_records,
        epms_records=epms_records,
        pr_model_rows=canonical.pr_model_rows,
        metadata=canonical.metadata,
    )


def epms_matcher(dataset: CanonicalDataset) -> BusinessDataset:
    epms_by_du: Dict[str, EPMSRecord] = {}
    for record in dataset.epms_records:
        du = record.canonical.get("du")
        if du and du not in epms_by_du:
            epms_by_du[du] = record

    records: List[BusinessRecord] = []
    for final_record in dataset.final_po_records:
        epms = epms_by_du.get(final_record.canonical.get("du"))
        scope = infer_scope(final_record)
        facts = build_business_facts(epms)
        expected_subcon = expected_subcontractor(scope, facts)
        records.append(
            BusinessRecord(
                final_po=final_record,
                epms=epms,
                scope=scope,
                expected_subcontractor=expected_subcon,
                business_facts=facts,
            )
        )
    return BusinessDataset(records=records, pr_model_rows=dataset.pr_model_rows, metadata=dataset.metadata)


def business_fact_builder(dataset: BusinessDataset) -> BusinessDataset:
    return dataset


def pr_model_resolver(dataset: BusinessDataset) -> BusinessDataset:
    return dataset


def expected_item_generator(dataset: BusinessDataset) -> ExpectedDataset:
    expected_records: List[ExpectedRecord] = []
    for record in dataset.records:
        items, evidence = expected_items_for_record(record, dataset.pr_model_rows)
        expected_records.append(
            ExpectedRecord(
                business_record=record,
                expected_items=items,
                expected_subcontractor=record.expected_subcontractor,
                model_evidence=evidence,
            )
        )
    return ExpectedDataset(records=expected_records, metadata=dataset.metadata)


def audit_engine(dataset: ExpectedDataset) -> AuditDataset:
    results: List[AuditResult] = []
    for expected_record in dataset.records:
        business = expected_record.business_record
        final_po = business.final_po
        submitted = final_po.canonical
        submitted_code = text(submitted.get("submitted_item_code"))
        submitted_subcon = normalize_subcontractor(submitted.get("submitted_subcontractor"))
        expected_subcon = normalize_subcontractor(expected_record.expected_subcontractor)
        epms_evidence = epms_evidence_text(business)

        if business.epms is None:
            results.append(
                make_result(
                    expected_record,
                    "Abnormal - Invalid PO",
                    "INVALID_NO_EXPECTED_BUSINESS_FACT",
                    "No EPMS row matched the submitted DU.",
                    epms_evidence,
                    consumes=False,
                )
            )
            continue

        if expected_subcon and submitted_subcon != expected_subcon:
            results.append(
                make_result(
                    expected_record,
                    "Abnormal - Invalid PO",
                    "INVALID_SUBCON_CHANGED",
                    "Submitted subcontractor does not match expected subcontractor.",
                    epms_evidence,
                    consumes=False,
                )
            )
            continue

        if business.scope == "UNKNOWN":
            results.append(
                make_result(
                    expected_record,
                    "Abnormal - Invalid PO",
                    "INVALID_WRONG_DOMAIN",
                    "Submitted item/domain could not be mapped to a supported TX audit scope.",
                    epms_evidence,
                    consumes=False,
                )
            )
            continue

        if business.scope == "OPERATION" and not business.business_facts.get("integration_end_date"):
            results.append(
                make_result(
                    expected_record,
                    "Abnormal - Invalid PO",
                    "INVALID_NO_OPERATION_TRIGGER",
                    "Operation item submitted but EPMS integration end date trigger is blank.",
                    epms_evidence,
                    consumes=False,
                )
            )
            continue

        if not expected_record.expected_items:
            results.append(
                make_result(
                    expected_record,
                    "Abnormal - Invalid PO",
                    "INVALID_NO_EXPECTED_BUSINESS_FACT",
                    "No expected item could be generated from current EPMS and PR Model inputs.",
                    epms_evidence,
                    consumes=False,
                )
            )
            continue

        expected_codes = {item.code for item in expected_record.expected_items}
        if submitted_code not in expected_codes:
            reason = wrong_reason_for_submitted_code(business, submitted_code)
            results.append(
                make_result(
                    expected_record,
                    "Abnormal - Wrong PO",
                    reason,
                    "Submitted item is in the auditable scope but does not match expected mapping.",
                    epms_evidence,
                    consumes=False,
                )
            )
            continue

        expected_quantity = sum(item.quantity for item in expected_record.expected_items if item.code == submitted_code)
        results.append(
            make_result(
                expected_record,
                "PENDING_QUANTITY",
                "PENDING_QUANTITY",
                "Submitted item passed validity checks and is ready for duplicate resolution.",
                epms_evidence,
                expected_quantity=expected_quantity,
                consumes=True,
            )
        )
    return AuditDataset(results=results, metadata=dataset.metadata)


def duplicate_resolver(dataset: AuditDataset) -> AuditDataset:
    results = list(dataset.results)
    pending = [idx for idx, result in enumerate(results) if result.consumes_quantity]
    pending.sort(key=lambda idx: results[idx].final_po.canonical.get("dispatch_sort_key"))
    consumed: Dict[Tuple[str, str, str], float] = defaultdict(float)

    for idx in pending:
        result = results[idx]
        data = result.final_po.canonical
        key = (
            text(data.get("du")),
            result.scope,
            text(data.get("submitted_item_code")),
        )
        submitted_qty = to_float(data.get("submitted_quantity"), default=0.0)
        remaining = max(result.expected_quantity - consumed[key], 0.0)
        normal_qty = min(submitted_qty, remaining)
        duplicate_qty = max(submitted_qty - normal_qty, 0.0)
        consumed[key] += normal_qty

        if duplicate_qty > 0 and normal_qty > 0:
            results[idx] = replace(
                result,
                classification="Abnormal - Duplicate PO",
                reason_code="DUPLICATE_PARTIAL_QUANTITY",
                normal_quantity=normal_qty,
                duplicate_quantity=duplicate_qty,
                explanation="Submitted quantity partially exceeds remaining expected quantity.",
                consumes_quantity=False,
            )
        elif duplicate_qty > 0:
            results[idx] = replace(
                result,
                classification="Abnormal - Duplicate PO",
                reason_code="DUPLICATE_FULL_QUANTITY",
                normal_quantity=0.0,
                duplicate_quantity=duplicate_qty,
                explanation="Expected quantity was already exhausted by earlier valid claims.",
                consumes_quantity=False,
            )
        else:
            reason_code = "NORMAL_FULL" if normal_qty == submitted_qty else "NORMAL_PARTIAL"
            results[idx] = replace(
                result,
                classification="Normal",
                reason_code=reason_code,
                normal_quantity=normal_qty,
                duplicate_quantity=0.0,
                explanation="Submitted claim matches expected item and available quantity.",
                consumes_quantity=False,
            )
    return AuditDataset(results=results, metadata=dataset.metadata)


def report_writer(dataset: AuditDataset, output_path: Path, summary_json: Optional[Path]) -> Dict[str, Any]:
    Workbook, _, Font, PatternFill = require_openpyxl()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "audit_result"

    source_headers = list(dataset.results[0].final_po.raw.keys()) if dataset.results else []
    source_headers = [h for h in source_headers if h != "_source_row"]
    audit_headers = [
        "Source Row",
        "Scope",
        "Audit Result",
        "Reason Code",
        "Expected Item",
        "Expected Quantity",
        "Expected Subcontractor",
        "Normal Quantity",
        "Duplicate Quantity",
        "EPMS Evidence",
        "PR Model Evidence",
        "Explanation",
    ]
    headers = source_headers + audit_headers
    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    header_font = Font(bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, result in enumerate(dataset.results, 2):
        row_values = [result.final_po.raw.get(header) for header in source_headers]
        row_values += [
            result.final_po.source_row,
            result.scope,
            result.classification,
            result.reason_code,
            "; ".join(item.code for item in result.expected_items),
            result.expected_quantity,
            result.expected_subcontractor,
            result.normal_quantity,
            result.duplicate_quantity,
            result.epms_evidence,
            result.pr_model_evidence,
            result.explanation,
        ]
        for col_idx, value in enumerate(row_values, 1):
            ws.cell(row_idx, col_idx, value)

    ws.freeze_panes = "A2"
    for idx, header in enumerate(headers, 1):
        width = min(max(len(str(header)) + 2, 12), 40)
        ws.column_dimensions[column_name(idx)].width = width

    wb.save(output_path)

    summary = {
        "output": str(output_path),
        "total_rows": len(dataset.results),
        "classifications": dict(Counter(result.classification for result in dataset.results)),
        "reason_codes": dict(Counter(result.reason_code for result in dataset.results)),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }
    if summary_json:
        summary_json.parent.mkdir(parents=True, exist_ok=True)
        summary_json.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    return summary


def run_pipeline(args: argparse.Namespace) -> Dict[str, Any]:
    final_po = require_file(Path(args.final_po), "Final PO")
    epms = require_file(Path(args.epms), "EPMS")
    pr_model = require_file(Path(args.pr_model), "PR Model")
    raw = workbook_reader(
        final_po,
        epms,
        pr_model,
        final_po_sheet=args.final_po_sheet,
        epms_sheet=args.epms_sheet,
    )
    mapped = field_mapper(raw)
    canonical = canonical_builder(mapped)
    matched = epms_matcher(canonical)
    business = business_fact_builder(matched)
    resolved = pr_model_resolver(business)
    expected = expected_item_generator(resolved)
    audited = audit_engine(expected)
    duplicated = duplicate_resolver(audited)
    return report_writer(
        duplicated,
        Path(args.output),
        Path(args.summary_json) if args.summary_json else None,
    )


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def first_nonempty(values: Iterable[Any]) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return None


def to_float(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value))
    return float(match.group(0)) if match else default


def parse_antenna_size(value: Any) -> Optional[float]:
    sizes = parse_antenna_sizes(value)
    return max(sizes) if sizes else None


def parse_antenna_sizes(value: Any) -> List[float]:
    raw = text(value).replace(",", ".")
    if not raw:
        return []
    sizes: List[float] = []
    for match in re.findall(r"(\d+(?:\.\d+)?)", raw):
        size = float(match)
        if 0 < size <= 5.0:
            sizes.append(size)
    return sorted(set(sizes))


def determine_ti_chosen_antenna_size(ne_size: Any, fe_size: Any) -> Optional[float]:
    sizes = [size for size in [parse_antenna_size(ne_size), parse_antenna_size(fe_size)] if size is not None]
    return max(sizes) if sizes else None


def expected_antenna_code(ne_size: Any, fe_size: Any) -> Optional[str]:
    sizes = [size for size in [parse_antenna_size(ne_size), parse_antenna_size(fe_size)] if size is not None]
    if not sizes:
        return None
    size = max(sizes)
    if size <= 0.6:
        return "350001095405"
    if size <= 1.2:
        return "350001095406"
    if size <= 1.8:
        return "350001095407"
    return "350001095408"


def normalize_choice_category(value: Any) -> Optional[str]:
    lower = text(value).lower()
    if not lower:
        return None
    if "simple packing" in lower:
        return "outbound_route"
    if "partial material transportation" in lower:
        return "material_route"
    if "antenna" in lower and parse_antenna_sizes(lower):
        return "antenna"
    if "inland transportation" in lower:
        return "inbound_route"
    if "dismantling" in lower and "antenna" in lower:
        return "antenna"
    return "choose"


def get_region_search_terms(region: Any) -> List[str]:
    region_key = text(region).lower()
    mapping = {
        "northern": ["north region", "perlis", "kedah", "penang", "perak"],
        "southern": ["south region", "negeri sembilan", "malacca", "johor"],
        "eastern": ["east region", "pahang", "terengganu", "kelantan"],
        "sabah": ["sabah"],
        "sarawak": ["sarawak", "salawak", "kuching", "sibu", "bintulu", "miri", "limbang", "lawas", "sri aman"],
        "central": ["kv region", "kv warehouse", "kuantan", "kk"],
    }
    return [term for term in mapping.get(region_key, [region_key]) if term]


def row_text(row: Dict[str, Any]) -> str:
    return " ".join([text(row.get("sow")), text(row.get("description")), text(row.get("rules"))])


def row_matches_chosen_size(row: Dict[str, Any], chosen_size: Optional[float]) -> bool:
    if chosen_size is None:
        return False
    return any(abs(size - chosen_size) < 1e-6 for size in parse_antenna_sizes(row_text(row)))


def is_mw_hardware_cutover_row(row: Dict[str, Any]) -> bool:
    return "mw hardware cutover" in row_text(row).lower()


def is_mw_reroute_sow(sow: Any) -> bool:
    lower = text(sow).lower()
    return "mw" in lower and "reroute" in lower


def classify_mw_reroute_model_row(row: Dict[str, Any]) -> str:
    lower = row_text(row).lower()
    if "new - mw link" in lower:
        return "install"
    if "mw dismantling" in lower and "antenna" in lower:
        return "dismantle"
    return "other"


def row_matches_size_bucket(row: Dict[str, Any], chosen_size: Optional[float]) -> bool:
    if chosen_size is None:
        return False
    normalized = row_text(row).replace(",", ".").lower()
    if ">3.2" in normalized:
        return chosen_size > 3.2
    for first, second in re.findall(r"(\d+(?:\.\d+)?)\s*[-/]\s*(\d+(?:\.\d+)?)\s*m?", normalized):
        low = float(first)
        high = float(second)
        if low <= chosen_size <= high:
            return True
    return row_matches_chosen_size(row, chosen_size)


def select_mw_reroute_row(rows: List[Dict[str, Any]], chosen_size: Optional[float]) -> Tuple[Optional[Dict[str, Any]], str]:
    matched = [row for row in rows if row_matches_size_bucket(row, chosen_size)]
    if len(matched) == 1:
        return matched[0], "matched"
    if len(matched) > 1:
        return None, "ambiguous"
    return None, "missing"


def match_mw_reroute_rows_like_create_pr_cd(
    facts: Dict[str, Any],
    ti_rows: List[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], List[str]]:
    candidates = [
        row
        for row in ti_rows
        if row.get("is_mandatory") and ("REROUTE" in text(row.get("sow")).upper() or text(row.get("sow")).upper() == "MW REROUTE")
    ]
    install_rows = [row for row in candidates if classify_mw_reroute_model_row(row) == "install"]
    dismantle_rows = [row for row in candidates if classify_mw_reroute_model_row(row) == "dismantle"]

    install_size = determine_ti_chosen_antenna_size(facts.get("antenna_size_ne"), facts.get("antenna_size_fe"))
    selected_rows: List[Dict[str, Any]] = []
    review_reasons: List[str] = []

    install_row, install_status = select_mw_reroute_row(install_rows, install_size)
    if install_row:
        selected_rows.append(install_row)
    else:
        review_reasons.append(f"MW Reroute install item not matched ({install_status})")

    # Dismantle antenna size detection in create-pr-cd uses contextual text
    # parsing. Until that is ported fully, do not guess a dismantle row.
    if dismantle_rows:
        review_reasons.append("MW Reroute dismantle item requires contextual size resolver")

    return selected_rows, review_reasons


def normalize_subcontractor(value: Any) -> str:
    raw = text(value).upper()
    raw = raw.replace("&", " AND ")
    raw = re.sub(r"\b(SDN|BHD|SDN\.|BHD\.|BERHAD|LTD|LIMITED|ENGINEERING|TECHNOLOGY|TECHNOLOGIES|SCIENCE|MALAYSIA|\\.|\\(|\\))\b", " ", raw)
    raw = re.sub(r"\s+", " ", raw).strip()
    if "GIROBUMI" in raw or raw == "GTSB":
        return "GTSB"
    if "GCI" in raw:
        return "GCI"
    if "ALL STAR" in raw or "ALLSTAR" in raw:
        return "ALLSTAR"
    if "SERIKANDI" in raw or "SERI PANCAR" in raw:
        return re.sub(r"\s+", " ", raw).strip()
    return raw


def infer_scope(record: FinalPORecord) -> str:
    data = record.canonical
    code = text(data.get("submitted_item_code"))
    domain = text(data.get("business_domain"))
    description = text(data.get("submitted_item_description"))
    combined = f"{domain} {description}".lower()
    if code == PLANNING_ITEM_CODE:
        return "PLANNING"
    if code == OPERATION_ITEM_CODE:
        return "OPERATION"
    if "勘察" in domain or "survey" in combined or "tss" in combined:
        return "TSS"
    if "微波" in domain or "installation" in combined or "antenna" in combined or code in ANTENNA_ITEM_CODES:
        return "TI"
    return "UNKNOWN"


def build_business_facts(epms: Optional[EPMSRecord]) -> Dict[str, Any]:
    if epms is None:
        return {}
    return dict(epms.canonical)


def expected_subcontractor(scope: str, facts: Dict[str, Any]) -> str:
    if scope == "TSS":
        return text(facts.get("expected_tss_subcontractor"))
    if scope == "TI":
        return text(facts.get("expected_ti_subcontractor"))
    if scope == "PLANNING":
        return text(facts.get("expected_planning_subcontractor"))
    if scope == "OPERATION":
        return "Allstar"
    return ""


def expected_items_for_record(record: BusinessRecord, pr_model_rows: List[Dict[str, Any]]) -> Tuple[List[ExpectedItem], str]:
    facts = record.business_facts
    scope = record.scope
    if record.epms is None:
        return [], "No EPMS match"
    if scope == "PLANNING":
        return [ExpectedItem(PLANNING_ITEM_CODE, 1.0, scope, "Confirmed Planning rule")], "Planning fixed rule"
    if scope == "OPERATION":
        if not facts.get("integration_end_date"):
            return [], "Operation trigger missing"
        return [ExpectedItem(OPERATION_ITEM_CODE, 1.0, scope, "Confirmed Operation rule")], "Operation fixed rule"
    if scope == "TSS":
        return expected_items_from_pr_model(scope, facts, pr_model_rows)
    if scope == "TI":
        return expected_ti_items_like_create_pr_cd(facts, pr_model_rows)
    return [], "Unsupported scope"


def expected_items_from_pr_model(scope: str, facts: Dict[str, Any], rows: List[Dict[str, Any]]) -> Tuple[List[ExpectedItem], str]:
    tx_sow = text(facts.get("tx_sow")).upper()
    if not tx_sow:
        return [], f"{scope}: missing EPMS Tx SOW"
    items: List[ExpectedItem] = []
    evidence_rows: List[str] = []
    for row in rows:
        if row.get("section") != scope:
            continue
        rules = text(row.get("rules")).upper()
        if "MANDATORY" not in rules:
            continue
        sow = text(row.get("sow")).upper()
        if sow and (sow in tx_sow or tx_sow in sow):
            items.append(
                ExpectedItem(
                    code=text(row.get("code")),
                    quantity=to_float(row.get("quantity"), default=1.0),
                    scope=scope,
                    reason=f"PR Model row {row.get('source_row')}",
                )
            )
            evidence_rows.append(f"{row.get('worksheet')}!{row.get('source_row')}")
    if not items:
        return [], f"{scope}: no mandatory PR Model match for Tx SOW={tx_sow}"
    return dedupe_expected_items(items), "; ".join(evidence_rows)


def expected_ti_items_like_create_pr_cd(facts: Dict[str, Any], rows: List[Dict[str, Any]]) -> Tuple[List[ExpectedItem], str]:
    tx_sow = text(facts.get("tx_sow"))
    if not tx_sow:
        return [], "TI: missing EPMS Tx SOW"

    ti_rows = [row for row in rows if row.get("section") == "TI"]
    chosen_size = determine_ti_chosen_antenna_size(
        facts.get("antenna_size_ne"),
        facts.get("antenna_size_fe"),
    )

    if is_mw_reroute_sow(tx_sow):
        selected_rows, review_reasons = match_mw_reroute_rows_like_create_pr_cd(facts, ti_rows)
        if review_reasons and not selected_rows:
            return [], "TI MW Reroute unresolved: " + "; ".join(review_reasons)
    else:
        selected_rows, review_required = match_ti_rows_like_create_pr_cd(
            tx_sow,
            chosen_size,
            text(facts.get("epms_region")),
            ti_rows,
        )
        if review_required and not selected_rows:
            return [], "TI choose-group unresolved using create-pr-cd selection rules"

    if not selected_rows:
        return [], f"TI: no mandatory PR Model match for Tx SOW={tx_sow}"

    items = [
        ExpectedItem(
            code=text(row.get("code")),
            quantity=to_float(row.get("quantity"), default=1.0),
            scope="TI",
            reason=f"create-pr-cd PR Model row {row.get('source_row')}",
        )
        for row in selected_rows
    ]
    evidence = "; ".join(f"{row.get('worksheet')}!{row.get('source_row')}" for row in selected_rows)
    return dedupe_expected_items(items), evidence


def match_ti_rows_like_create_pr_cd(
    sow: str,
    chosen_size: Optional[float],
    region: str,
    ti_rows: List[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], bool]:
    sow_upper = sow.upper()
    candidates: List[Dict[str, Any]] = []
    for row in ti_rows:
        item_sow_upper = text(row.get("sow")).upper()
        if item_sow_upper == sow_upper or item_sow_upper in sow_upper or sow_upper in item_sow_upper:
            if row.get("is_mandatory") and not is_mw_hardware_cutover_row(row):
                candidates.append(row)

    if not candidates:
        return [], False

    grouped_candidates: Dict[Tuple[str, ...], List[Dict[str, Any]]] = {}
    for row in candidates:
        rules = text(row.get("rules")).lower()
        if "choose" in rules:
            group_key = (
                text(row.get("sow")),
                text(row.get("rules")),
                text(normalize_choice_category(" ".join([text(row.get("sow")), text(row.get("description")), text(row.get("rules"))]))),
            )
        else:
            group_key = (text(row.get("code")),)
        grouped_candidates.setdefault(group_key, []).append(row)

    selected_rows: List[Dict[str, Any]] = []
    review_required = False
    choose_group_ambiguous = False
    for group_rows in grouped_candidates.values():
        if len(group_rows) == 1:
            selected_rows.extend(group_rows)
            continue

        rules = text(group_rows[0].get("rules")).lower()
        if "choose" in rules:
            chosen_rows, ambiguous = filter_choose_group_rows_like_create_pr_cd(group_rows, chosen_size, region)
            selected_rows.extend(chosen_rows)
            if ambiguous:
                review_required = True
                choose_group_ambiguous = True
        else:
            selected_rows.extend(group_rows)
            if len(group_rows) > 1:
                review_required = True

    if choose_group_ambiguous:
        return [], True

    return selected_rows, review_required


def filter_choose_group_rows_like_create_pr_cd(
    group_rows: List[Dict[str, Any]],
    chosen_size: Optional[float],
    region: str,
) -> Tuple[List[Dict[str, Any]], bool]:
    if len(group_rows) <= 1:
        return group_rows, False

    category = normalize_choice_category(
        " ".join([text(group_rows[0].get("sow")), text(group_rows[0].get("description")), text(group_rows[0].get("rules"))])
    )

    if category == "antenna":
        matched = [row for row in group_rows if row_matches_chosen_size(row, chosen_size)]
        return (matched, False) if len(matched) == 1 else ([], True)

    if category in {"material_route", "choose"}:
        for term in get_region_search_terms(region):
            matched = [
                row
                for row in group_rows
                if term in " ".join([text(row.get("sow")), text(row.get("description")), text(row.get("rules"))]).lower()
            ]
            if len(matched) == 1:
                return matched, False
        return [], True

    # create-pr-cd uses GeographyResolver for these route groups. The auditor
    # currently fails closed until that resolver is wired in.
    if category in {"outbound_route", "inbound_route"}:
        return [], True

    return [], True


def dedupe_expected_items(items: List[ExpectedItem]) -> List[ExpectedItem]:
    by_code: Dict[Tuple[str, str], ExpectedItem] = {}
    for item in items:
        key = (item.scope, item.code)
        if key in by_code:
            previous = by_code[key]
            by_code[key] = replace(previous, quantity=previous.quantity + item.quantity)
        else:
            by_code[key] = item
    return list(by_code.values())


def wrong_reason_for_submitted_code(business: BusinessRecord, submitted_code: str) -> str:
    if business.scope == "TI" and submitted_code in ANTENNA_ITEM_CODES:
        return "WRONG_ANTENNA_SIZE"
    if business.scope == "TI":
        return "WRONG_LINE_ITEM_MAPPING"
    if business.scope in {"TSS", "PLANNING", "OPERATION"}:
        return "WRONG_LINE_ITEM_MAPPING"
    return "INVALID_WRONG_DOMAIN"


def make_result(
    expected_record: ExpectedRecord,
    classification: str,
    reason_code: str,
    explanation: str,
    epms_evidence: str,
    expected_quantity: Optional[float] = None,
    consumes: bool = False,
) -> AuditResult:
    if expected_quantity is None:
        submitted_code = text(expected_record.business_record.final_po.canonical.get("submitted_item_code"))
        expected_quantity = sum(
            item.quantity for item in expected_record.expected_items if item.code == submitted_code
        )
    return AuditResult(
        final_po=expected_record.business_record.final_po,
        scope=expected_record.business_record.scope,
        classification=classification,
        reason_code=reason_code,
        expected_items=expected_record.expected_items,
        expected_subcontractor=expected_record.expected_subcontractor,
        expected_quantity=expected_quantity,
        normal_quantity=0.0,
        duplicate_quantity=0.0,
        explanation=explanation,
        epms_evidence=epms_evidence,
        pr_model_evidence=expected_record.model_evidence,
        consumes_quantity=consumes,
    )


def epms_evidence_text(record: BusinessRecord) -> str:
    if record.epms is None:
        return "No EPMS match"
    facts = record.business_facts
    return (
        f"EPMS row {record.epms.source_row}; "
        f"DU={facts.get('du')}; "
        f"Tx SOW={facts.get('tx_sow')}; "
        f"Region={facts.get('epms_region')}"
    )


def dispatch_sort_key(data: Dict[str, Any]) -> Tuple[Any, str, str, int]:
    return (
        sortable_date(data.get("dispatch_date")),
        text(data.get("request_number")),
        text(data.get("dispatch_order_number")),
        int(to_float(data.get("po_line_number"), default=0.0)),
    )


def sortable_date(value: Any) -> Any:
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        return value
    value_text = text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value_text, fmt)
        except ValueError:
            pass
    return value_text


def column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def main() -> int:
    args = parse_args()
    try:
        summary = run_pipeline(args)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print(json.dumps(summary, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
